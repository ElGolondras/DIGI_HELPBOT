import customtkinter as ctk
from PIL import Image, ImageDraw
from groq import Groq
import threading
import time
import os
import json
import vlc
from datetime import datetime
import base64
from tkinter import filedialog
import mysql.connector
import fitz          # pymupdf
from docx import Document as DocxDocument
import openpyxl
import tempfile
import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io

def recurso_path(relative_path):
    import sys
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ─────────────────────────────────────────────
#  CARGA DE VARIABLES DE ENTORNO (.env)
# ─────────────────────────────────────────────
def _cargar_env():
    env_path = recurso_path(".env") 
    if os.path.exists(env_path):
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())

_cargar_env()

API_KEY = os.environ.get("GROQ_API_KEY", "")  # Define GROQ_API_KEY en tu .env

# ─────────────────────────────────────────────
#  CONFIGURACIÓN GOOGLE DRIVE (SERVICE ACCOUNT)
# ─────────────────────────────────────────────
DRIVE_SERVICE_ACCOUNT_FILE = recurso_path(
    os.environ.get("DRIVE_SERVICE_ACCOUNT_FILE", "service_account.json")
)
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

def _get_drive_service():
    """Devuelve un cliente autenticado de Google Drive."""
    creds = service_account.Credentials.from_service_account_file(
        DRIVE_SERVICE_ACCOUNT_FILE, scopes=DRIVE_SCOPES
    )
    return build("drive", "v3", credentials=creds)

def descargar_video_drive(drive_id):
    """
    Descarga un video de Google Drive por su file_id a un archivo temporal.
    Devuelve la ruta del temporal o None si falla.

    El 'drive_id' es el ID que aparece en la URL de Drive:
        https://drive.google.com/file/d/ESTE_ES_EL_ID/view
    """
    try:
        service  = _get_drive_service()
        meta     = service.files().get(fileId=drive_id, fields="name,mimeType").execute()
        nombre   = meta.get("name", "video.mp4")
        ext      = os.path.splitext(nombre)[1] or ".mp4"

        request  = service.files().get_media(fileId=drive_id)
        buf      = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request, chunksize=8 * 1024 * 1024)

        done = False
        while not done:
            _, done = downloader.next_chunk()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        tmp.write(buf.getvalue())
        tmp.close()
        print(f"[Drive] Video descargado → {tmp.name} ({len(buf.getvalue())/1024/1024:.1f} MB)")
        return tmp.name
    except Exception as e:
        print(f"[Drive] Error descargando video: {e}")
        return None

# ─────────────────────────────────────────────
#  CONFIGURACIÓN MySQL
# ─────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.environ.get("DB_HOST", "localhost"),
    "port":     int(os.environ.get("DB_PORT", 3306)),
    "user":     os.environ.get("DB_USER", "root"),
    "password": os.environ.get("DB_PASSWORD", ""),
    "database": os.environ.get("DB_NAME", "digihelp"),
}

def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)

def login_usuario(username, contrasenia):
    """Devuelve el dict del usuario si las credenciales son correctas, si no None."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM usuarios WHERE username = %s AND contrasenia = %s",
            (username, contrasenia)
        )
        usuario = cursor.fetchone()
        cursor.close()
        conn.close()
        return usuario  # dict con id, username, nombre_completo, departamento, email, contrasenia
    except Exception as e:
        print(f"[DB] Error en login: {e}")
        return None

def detectar_urgencia(problema):
    """Detecta automáticamente la urgencia según el texto del problema."""
    pl = problema.lower()
    if any(p in pl for p in ["no enciende","pantalla azul","virus","hackeado","datos perdidos",
                              "no arranca","caido","servidor","ransomware","brecha","seguridad"]):
        return "alta"
    elif any(p in pl for p in ["impresora","internet","red","correo","contrasena","contraseña",
                                "vpn","lento","cuelga","no conecta","wifi","actualizar"]):
        return "media"
    return "baja"

def crear_incidencia_db(nombre_completo, departamento, email, problema, urgencia="media"):
    """Crea una incidencia con todos los datos del usuario."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """INSERT INTO incidencias (usuario, departamento, email, problema, urgencia)
               VALUES (%s, %s, %s, %s, %s)""",
            (nombre_completo, departamento, email, problema, urgencia)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error creando incidencia: {e}")
        return False

# ─────────────────────────────────────────────
#  CHATS EN MySQL POR USUARIO
# ─────────────────────────────────────────────
def guardar_chat_db(usuario_id, chat_id, titulo, fecha, mensajes_json):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO chats (usuario_id, chat_id, titulo, fecha, mensajes)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE titulo=%s, fecha=%s, mensajes=%s
        """, (usuario_id, chat_id, titulo, fecha, mensajes_json,
              titulo, fecha, mensajes_json))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error guardando chat: {e}")
        return False

def cargar_chats_db(usuario_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT chat_id, titulo, fecha, mensajes FROM chats WHERE usuario_id=%s ORDER BY fecha DESC",
            (usuario_id,)
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows
    except Exception as e:
        print(f"[DB] Error cargando chats: {e}")
        return []

def eliminar_chat_db(usuario_id, chat_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM chats WHERE usuario_id=%s AND chat_id=%s", (usuario_id, chat_id))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"[DB] Error eliminando chat: {e}")

# ─────────────────────────────────────────────
#  PREFERENCIAS EN MySQL POR USUARIO
# ─────────────────────────────────────────────
def guardar_prefs_db(usuario_id, prefs):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        prefs_json = json.dumps(prefs, ensure_ascii=False)
        cursor.execute("""
            INSERT INTO preferencias (usuario_id, prefs)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE prefs=%s
        """, (usuario_id, prefs_json, prefs_json))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"[DB] Error guardando prefs: {e}")

def cargar_prefs_db(usuario_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT prefs FROM preferencias WHERE usuario_id=%s", (usuario_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if row:
            return json.loads(row["prefs"])
    except Exception as e:
        print(f"[DB] Error cargando prefs: {e}")
    return {"tema": "claro", "acento": "Azul", "fuente": 13}
client = Groq(api_key=API_KEY)
MODELO_TEXTO  = "llama-3.3-70b-versatile"
MODELO_VISION = "meta-llama/llama-4-scout-17b-16e-instruct"

SYSTEM_PROMPT = """Eres DigiHelp, asistente de soporte técnico IT. Tu única función es resolver problemas técnicos.

INSTRUCCIÓN PRINCIPAL: SIEMPRE saluda al principio y haz nada mas que un saludo. Luego responde directamente al problema.

EJEMPLOS DE CÓMO DEBES RESPONDER:

Usuario: "la impresora no me va"
TÚ: "No te preocupes, vamos a arreglarlo juntos. Sigue estos pasos:
1. Apaga la impresora — busca el botón de encendido y mantenlo pulsado hasta que se apague.
2. Desenchufa el cable de la luz — el que va a la pared.
3. Espera 30 segundos.
4. Vuelve a enchufarlo y enciende la impresora.
¿Ha vuelto a funcionar?"

Usuario: "no tengo internet"
TÚ: "Vamos a solucionarlo paso a paso:
1. Mira el router — esa cajita con lucecitas que da el wifi. ¿Las luces están encendidas?
2. Si alguna luz está roja o apagada, apaga el router (botón detrás) y vuelve a encenderlo.
3. Espera 1 minuto y prueba de nuevo.
¿Ha funcionado?"

REGLAS:
- En la primera conversacion empieza con un saludo (¡Hola!, ¡Buenos días!, ¡Buenas tardes!, etc.).
- Responde directamente al problema sin rodeos.
- Si hay una imagen analízala describiendo qué ves.
- Si la pregunta no es de IT (recetas, noticias...) responde: "Solo puedo ayudarte con problemas técnicos. ¿Tienes alguna incidencia IT?"
- Usa lenguaje muy simple, como si hablaras con alguien que nunca ha usado un ordenador.
- Pasos cortos, máximo 2 líneas cada uno.
- Siempre pregunta al final si ha funcionado."""

C = {
    "bg_app":        "#F0F2F5",
    "sidebar_bg":    "#1B2A4A",
    "sidebar_hover": "#243557",
    "sidebar_btn":   "#2C3E6B",
    "accent":        "#2563EB",
    "accent_dark":   "#1D4ED8",
    "bubble_user":   "#2563EB",
    "bubble_ia":     "#FFFFFF",
    "text_dark":     "#1E293B",
    "text_light":    "#FFFFFF",
    "text_muted":    "#94A3B8",
    "input_bg":      "#FFFFFF",
    "border":        "#CBD5E1",
    "card":          "#FFFFFF",
    "header_bg":     "#FFFFFF",
    "video_bg":      "#0F172A",
}
TEMAS = {
    "claro": {
        "bg_app": "#F0F2F5", "sidebar_bg": "#1B2A4A", "sidebar_hover": "#243557",
        "sidebar_btn": "#2C3E6B", "bubble_ia": "#FFFFFF", "text_dark": "#1E293B",
        "text_light": "#FFFFFF", "text_muted": "#94A3B8", "input_bg": "#FFFFFF",
        "border": "#CBD5E1", "card": "#FFFFFF", "header_bg": "#FFFFFF", "video_bg": "#0F172A",
    },
    "oscuro": {
        "bg_app": "#0F172A", "sidebar_bg": "#0A0F1E", "sidebar_hover": "#1E293B",
        "sidebar_btn": "#1E293B", "bubble_ia": "#1E293B", "text_dark": "#F1F5F9",
        "text_light": "#FFFFFF", "text_muted": "#64748B", "input_bg": "#1E293B",
        "border": "#334155", "card": "#1E293B", "header_bg": "#0F172A", "video_bg": "#000000",
    },
}

ACENTOS = {
    "Azul":    ("#2563EB", "#1D4ED8"),
    "Violeta": ("#7C3AED", "#6D28D9"),
    "Verde":   ("#16A34A", "#15803D"),
    "Rojo":    ("#DC2626", "#B91C1C"),
    "Naranja": ("#EA580C", "#C2410C"),
    "Rosa":    ("#DB2777", "#BE185D"),
}

PREFS_PATH = os.path.join(os.path.abspath("."), "preferencias.json")

def cargar_prefs():
    try:
        with open(PREFS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"tema": "claro", "acento": "Azul", "fuente": 13}

def guardar_prefs(prefs):
    try:
        with open(PREFS_PATH, "w", encoding="utf-8") as f:
            json.dump(prefs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def aplicar_tema(prefs):
    """Actualiza el dict global C con el tema y acento seleccionados."""
    global FUENTE
    tema = TEMAS.get(prefs.get("tema", "claro"), TEMAS["claro"])
    acento_nombre = prefs.get("acento", "Azul")
    acento, acento_dark = ACENTOS.get(acento_nombre, ACENTOS["Azul"])
    C.update(tema)
    C["accent"]      = acento
    C["accent_dark"] = acento_dark
    C["bubble_user"] = acento
    FUENTE = int(prefs.get("fuente", 13))
    modo_ctk = "dark" if prefs.get("tema") == "oscuro" else "light"
    ctk.set_appearance_mode(modo_ctk)

FUENTE = 13
# Cargar y aplicar preferencias al inicio
_prefs_globales = cargar_prefs()
aplicar_tema(_prefs_globales)

def _extraer_drive_id(valor):
    """
    Extrae el file ID de Google Drive de un valor que puede ser:
      - El ID directamente:   1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs
      - URL completa:         https://drive.google.com/file/d/ID/view
      - URL de descarga:      https://drive.google.com/uc?id=ID
    Devuelve el ID si lo detecta, o None si parece una ruta local.
    """
    if not valor:
        return None
    # URL tipo /file/d/ID/
    if "drive.google.com/file/d/" in valor:
        try:
            return valor.split("/file/d/")[1].split("/")[0].split("?")[0]
        except Exception:
            return None
    # URL tipo ?id=ID
    if "drive.google.com" in valor and "id=" in valor:
        try:
            return valor.split("id=")[1].split("&")[0]
        except Exception:
            return None
    # Si no tiene separadores de ruta ni extensión de video, asumimos que es un ID directo
    extensiones_video = (".mp4", ".avi", ".mkv", ".mov", ".wmv", ".flv", ".webm")
    tiene_extension = any(valor.lower().endswith(e) for e in extensiones_video)
    tiene_separador = os.sep in valor or "/" in valor or "\\" in valor
    if not tiene_extension and not tiene_separador and len(valor) > 10:
        return valor  # parece un ID de Drive
    return None


class BurbujaChat(ctk.CTkFrame):
    def __init__(self, parent, texto, es_ia, avatar_ia=None, timestamp="", imagen_ruta=None, animar=False):
        super().__init__(parent, fg_color="transparent")
        self.es_ia = es_ia
        self._animar_entrada = animar
        self.pack(fill="x", padx=16, pady=6)
        if es_ia:
            self._burbuja_ia(texto, avatar_ia, timestamp)
        else:
            self._burbuja_usuario(texto, timestamp, imagen_ruta)
        if animar:
            self._fade_in()

    def _fade_in(self):
        """Efecto de deslizamiento suave al aparecer."""
        try:
            self.configure(fg_color="transparent")
            self._alpha_step = 0
            def _step():
                self._alpha_step += 1
                if self._alpha_step <= 8:
                    self.after(18, _step)
            _step()
        except Exception:
            pass

    def iniciar_puntos(self):
        """Animación de puntos ⋯ mientras la IA está escribiendo + pulso del avatar."""
        self._puntos_activos = True
        self._puntos_estado  = 0
        self._animar_puntos()
        self._pulso_activo = True
        self._animar_pulso()

    def detener_puntos(self):
        self._puntos_activos = False
        self._pulso_activo   = False
        # Restaurar avatar al color normal
        try:
            self._avatar_frame.configure(fg_color="transparent")
        except Exception:
            pass

    def _animar_puntos(self):
        if not self._puntos_activos:
            return
        estados = ["⏳ Escribiendo", "⏳ Escribiendo·", "⏳ Escribiendo··", "⏳ Escribiendo···"]
        try:
            self.lbl.configure(text=estados[self._puntos_estado % 4])
            self._puntos_estado += 1
            self.after(380, self._animar_puntos)
        except Exception:
            pass

    def _animar_pulso(self):
        if not self._pulso_activo:
            return
        try:
            colores = [C["accent"], C["accent_dark"], C["accent"]]
            color = colores[self._puntos_estado % len(colores)]
            self._avatar_frame.configure(fg_color=color)
            self.after(500, self._animar_pulso)
        except Exception:
            pass

    def recolorear(self):
        """Actualiza los colores de esta burbuja con los valores actuales de C y FUENTE."""
        try:
            if self.es_ia:
                self._frame_burbuja.configure(fg_color=C["bubble_ia"], border_color=C["border"])
                self.lbl.configure(text_color=C["text_dark"], font=("Helvetica", FUENTE))
                if hasattr(self, "_lbl_nombre"):
                    self._lbl_nombre.configure(text_color=C["accent"])
                if hasattr(self, "_lbl_ts"):
                    self._lbl_ts.configure(text_color=C["text_muted"])
            else:
                self._frame_burbuja.configure(fg_color=C["bubble_user"])
                self.lbl.configure(font=("Helvetica", FUENTE))
                if hasattr(self, "_lbl_ts"):
                    self._lbl_ts.configure(text_color=C["text_muted"])
        except Exception:
            pass

    def _burbuja_ia(self, texto, avatar, timestamp):
        fila = ctk.CTkFrame(self, fg_color="transparent")
        fila.pack(fill="x", anchor="w")
        self._avatar_frame = ctk.CTkFrame(fila, width=36, height=36, corner_radius=18, fg_color="transparent")
        self._avatar_frame.pack(side="left", anchor="n", padx=(0, 10), pady=2)
        self._avatar_frame.pack_propagate(False)
        if avatar:
            ctk.CTkLabel(self._avatar_frame, image=avatar, text="").pack(expand=True)
        else:
            try:
                img = Image.open(recurso_path("avatar.png")).convert("RGBA")
                s = min(img.size)
                img = img.crop(((img.width-s)//2, (img.height-s)//2, (img.width+s)//2, (img.height+s)//2))
                img = img.resize((36, 36), Image.LANCZOS)
                mask = Image.new("L", (36, 36), 0)
                ImageDraw.Draw(mask).ellipse((0, 0, 36, 36), fill=255)
                out = Image.new("RGBA", (36, 36), (0,0,0,0))
                out.paste(img, (0, 0), mask)
                ctk_img = ctk.CTkImage(light_image=out, dark_image=out, size=(36, 36))
                ctk.CTkLabel(self._avatar_frame, image=ctk_img, text="").pack(expand=True)
                self._avatar_frame._img_ref = ctk_img
            except Exception:
                ctk.CTkLabel(self._avatar_frame, text="D", font=("Georgia", 14, "bold"), text_color="white").pack(expand=True)
        contenido = ctk.CTkFrame(fila, fg_color="transparent")
        contenido.pack(side="left", fill="x", expand=True)
        cabecera = ctk.CTkFrame(contenido, fg_color="transparent")
        cabecera.pack(fill="x", anchor="w")
        self._lbl_nombre = ctk.CTkLabel(cabecera, text="DigiHelp AI", font=("Helvetica", 11, "bold"), text_color=C["accent"])
        self._lbl_nombre.pack(side="left")
        self._lbl_ts = ctk.CTkLabel(cabecera, text=f"  {timestamp}", font=("Helvetica", 10), text_color=C["text_muted"])
        self._lbl_ts.pack(side="left")
        self._frame_burbuja = ctk.CTkFrame(contenido, fg_color=C["bubble_ia"], corner_radius=12, border_width=1, border_color=C["border"])
        self._frame_burbuja.pack(fill="x", anchor="w", pady=(4, 0))
        self.lbl = ctk.CTkLabel(self._frame_burbuja, text=texto, font=("Helvetica", FUENTE), text_color=C["text_dark"],
                                wraplength=520, justify="left", padx=14, pady=12, anchor="w")
        self.lbl.pack(fill="x")

    def _burbuja_usuario(self, texto, timestamp, imagen_ruta=None):
        fila = ctk.CTkFrame(self, fg_color="transparent")
        fila.pack(fill="x", anchor="e")
        contenido = ctk.CTkFrame(fila, fg_color="transparent")
        contenido.pack(side="right")

        cabecera = ctk.CTkFrame(contenido, fg_color="transparent")
        cabecera.pack(fill="x", anchor="e")
        self._lbl_ts = ctk.CTkLabel(cabecera, text=f"{timestamp}  ", font=("Helvetica", 10), text_color=C["text_muted"])
        self._lbl_ts.pack(side="right")
        ctk.CTkLabel(cabecera, text="Tú", font=("Helvetica", 11, "bold"), text_color=C["text_muted"]).pack(side="right")

        self._frame_burbuja = ctk.CTkFrame(contenido, fg_color=C["bubble_user"], corner_radius=12)
        self._frame_burbuja.pack(anchor="e", pady=(4, 0))

        # Miniatura de imagen si existe
        if imagen_ruta and os.path.exists(imagen_ruta):
            try:
                img = Image.open(imagen_ruta).convert("RGB")
                img.thumbnail((220, 160), Image.LANCZOS)
                ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
                lbl_img = ctk.CTkLabel(self._frame_burbuja, image=ctk_img, text="", cursor="hand2")
                lbl_img.image = ctk_img
                lbl_img.pack(padx=10, pady=(10, 4))
            except Exception:
                pass

        # Texto (solo si hay algo que mostrar)
        texto_limpio = texto.replace("  🖼️", "").strip()
        if texto_limpio:
            self.lbl = ctk.CTkLabel(self._frame_burbuja, text=texto_limpio, font=("Helvetica", FUENTE),
                                    text_color=C["text_light"], wraplength=400,
                                    justify="right", padx=14, pady=10)
            self.lbl.pack()
        else:
            self.lbl = ctk.CTkLabel(self._frame_burbuja, text="", font=("Helvetica", 1))
            self.lbl.pack()

    def actualizar_texto(self, texto):
        self.lbl.configure(text=texto)


class LoginApp(ctk.CTk):
    """Ventana de login independiente. Al autenticarse lanza DigiHelpApp y se cierra."""
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        self.title("DigiHelp AI — Iniciar sesión")
        self.geometry("420x500")
        self.resizable(False, False)
        self.configure(fg_color=C["bg_app"])

        self.update_idletasks()
        x = (self.winfo_screenwidth()  - 420) // 2
        y = (self.winfo_screenheight() - 500) // 2
        self.geometry(f"420x500+{x}+{y}")

        self.protocol("WM_DELETE_WINDOW", self._cancelar)
        self._construir_ui()

    def _construir_ui(self):
        header = ctk.CTkFrame(self, fg_color=C["sidebar_bg"], corner_radius=0, height=120)
        header.pack(fill="x")
        header.pack_propagate(False)

        try:
            img = Image.open(recurso_path("avatar.png")).convert("RGBA")
            s = min(img.size)
            img = img.crop(((img.width-s)//2, (img.height-s)//2, (img.width+s)//2, (img.height+s)//2))
            img = img.resize((56, 56), Image.LANCZOS)
            mask = Image.new("L", (56, 56), 0)
            ImageDraw.Draw(mask).ellipse((0, 0, 56, 56), fill=255)
            out = Image.new("RGBA", (56, 56), (0,0,0,0))
            out.paste(img, (0, 0), mask)
            ctk_logo = ctk.CTkImage(light_image=out, dark_image=out, size=(56, 56))
            lbl_logo = ctk.CTkLabel(header, image=ctk_logo, text="")
            lbl_logo.pack(pady=(20, 4))
            lbl_logo._img_ref = ctk_logo
        except Exception:
            ctk.CTkLabel(header, text="D", font=("Georgia", 28, "bold"),
                         text_color="white").pack(pady=(20, 4))

        ctk.CTkLabel(header, text="DigiHelp AI", font=("Helvetica", 18, "bold"),
                     text_color="white").pack()

        form = ctk.CTkFrame(self, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=40, pady=30)

        ctk.CTkLabel(form, text="Usuario", font=("Helvetica", 13, "bold"),
                     text_color=C["text_dark"], anchor="w").pack(fill="x")
        self.entry_user = ctk.CTkEntry(form, placeholder_text="Introduce tu usuario",
                                       height=44, corner_radius=10,
                                       font=("Helvetica", 13))
        self.entry_user.pack(fill="x", pady=(4, 16))
        self.entry_user.bind("<Return>", lambda e: self.entry_pass.focus())

        ctk.CTkLabel(form, text="Contraseña", font=("Helvetica", 13, "bold"),
                     text_color=C["text_dark"], anchor="w").pack(fill="x")
        self.entry_pass = ctk.CTkEntry(form, placeholder_text="Introduce tu contraseña",
                                       height=44, corner_radius=10,
                                       font=("Helvetica", 13), show="•")
        self.entry_pass.pack(fill="x", pady=(4, 8))
        self.entry_pass.bind("<Return>", lambda e: self._intentar_login())

        self.lbl_error = ctk.CTkLabel(form, text="", font=("Helvetica", 12),
                                      text_color="#EF4444")
        self.lbl_error.pack(fill="x", pady=(0, 12))

        self.btn_login = ctk.CTkButton(form, text="Entrar", height=46, corner_radius=10,
                                       fg_color=C["accent"], hover_color=C["accent_dark"],
                                       font=("Helvetica", 14, "bold"), text_color="white",
                                       command=self._intentar_login)
        self.btn_login.pack(fill="x")

        ctk.CTkLabel(form, text="Soporte IT · v2.0", font=("Helvetica", 11),
                     text_color=C["text_muted"]).pack(pady=(20, 0))

        self.entry_user.focus()

    def _intentar_login(self):
        username = self.entry_user.get().strip()
        contrasenia = self.entry_pass.get().strip()
        if not username or not contrasenia:
            self.lbl_error.configure(text="Por favor, rellena todos los campos.")
            return
        self.btn_login.configure(state="disabled", text="Verificando...")
        self.lbl_error.configure(text="")

        def _verificar():
            usuario = login_usuario(username, contrasenia)
            if usuario:
                self.after(0, lambda: self._login_ok(usuario))
            else:
                self.after(0, lambda: self._login_fail())

        threading.Thread(target=_verificar, daemon=True).start()

    def _login_ok(self, usuario):
        self.destroy()
        app = DigiHelpApp(usuario)
        app.mainloop()

    def _login_fail(self):
        self.lbl_error.configure(text="Usuario o contraseña incorrectos.")
        self.btn_login.configure(state="normal", text="Entrar")
        self.entry_pass.delete(0, "end")
        self.entry_pass.focus()

    def _cancelar(self):
        self.destroy()
        import sys
        sys.exit(0)

class DigiHelpApp(ctk.CTk):
    def __init__(self, usuario_data):
        super().__init__()
        self.usuario_data = usuario_data  # dict con datos del usuario logueado
        self._usuario_id  = usuario_data.get("id")

        # Cargar y aplicar preferencias del usuario desde DB
        prefs_usuario = cargar_prefs_db(self._usuario_id)
        aplicar_tema(prefs_usuario)

        ctk.set_appearance_mode("dark" if prefs_usuario.get("tema") == "oscuro" else "light")
        self.title("DigiHelp AI — Soporte IT Corporativo")
        self.geometry("1200x800")
        self.minsize(900, 600)
        self.configure(fg_color=C["bg_app"])

        # Icono de la ventana y barra de tareas
        try:
            ruta_ico = recurso_path("avatar.ico")
            if os.path.exists(ruta_ico):
                self.iconbitmap(ruta_ico)
        except Exception:
            pass

        self._imagen_adjunta = None  # (ruta, base64_data, media_type)
        self._doc_adjunto = None      # (ruta, texto_extraido)
        self._intentos = 0            # Contador de intentos de resolución
        self._problema_inicial = None # Primer mensaje del usuario
        self._esperando_confirmacion = False  # Esperando si se resolvió
        self.historial = [{"role": "system", "content": SYSTEM_PROMPT}]
        self.chat_id = f"chat_{int(time.time())}"
        self._chat_titulo = None  # Se genera en el primer mensaje y se reutiliza
        self._chat_archivo = None  # Nombre del archivo JSON de este chat
        self.sidebar_visible = True
        self.video_cap = None   # ya no se usa, se mantiene por compatibilidad con cerrar_sesion
        self.video_activo = False
        self.welcome = None
        self._vlc_instance = None
        self._vlc_player   = None
        self._video_temp   = None   # archivo temporal descargado de Drive

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._construir_sidebar()
        self._construir_main()
        self.cargar_sidebar()

    def _construir_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=270, corner_radius=0, fg_color=C["sidebar_bg"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(2, weight=1)
        self.sidebar.grid_columnconfigure(0, weight=1)

        logo_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent", height=72)
        logo_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=(20, 0))
        logo_frame.grid_propagate(False)
        self._icono_frame = ctk.CTkFrame(logo_frame, width=38, height=38, corner_radius=19, fg_color=C["accent"])
        self._icono_frame.pack(side="left", padx=(0, 10))
        self._icono_frame.pack_propagate(False)
        try:
            img = Image.open(recurso_path("avatar.png")).convert("RGBA")
            s = min(img.size)
            img = img.crop(((img.width-s)//2, (img.height-s)//2, (img.width+s)//2, (img.height+s)//2))
            img = img.resize((38, 38), Image.LANCZOS)
            mask = Image.new("L", (38, 38), 0)
            ImageDraw.Draw(mask).ellipse((0, 0, 38, 38), fill=255)
            out = Image.new("RGBA", (38, 38), (0,0,0,0))
            out.paste(img, (0, 0), mask)
            ctk_logo = ctk.CTkImage(light_image=out, dark_image=out, size=(38, 38))
            ctk.CTkLabel(self._icono_frame, image=ctk_logo, text="").pack(expand=True)
            self._icono_frame._img_ref = ctk_logo
        except Exception:
            ctk.CTkLabel(self._icono_frame, text="D", font=("Georgia", 18, "bold"), text_color="white").pack(expand=True)
        ctk.CTkLabel(logo_frame, text="DigiHelp AI", font=("Helvetica", 17, "bold"), text_color="white").pack(side="left", anchor="w")

        self.btn_nuevo = ctk.CTkButton(self.sidebar, text="＋  Nuevo chat", font=("Helvetica", 13, "bold"),
            height=42, corner_radius=10, fg_color=C["accent"], hover_color=C["accent_dark"],
            text_color="white", command=self.nuevo_chat)
        self.btn_nuevo.grid(row=1, column=0, sticky="ew", padx=16, pady=(16, 8))

        self.lista_frame = ctk.CTkScrollableFrame(self.sidebar, fg_color="transparent",
                                                   scrollbar_button_color=C["sidebar_hover"])
        self.lista_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=4)

        footer = ctk.CTkFrame(self.sidebar, fg_color="transparent", corner_radius=0)
        footer.grid(row=3, column=0, sticky="ew", padx=12, pady=12)

        ctk.CTkLabel(footer, text="Soporte IT · v2.0", font=("Helvetica", 11),
                     text_color=C["text_muted"]).pack(pady=(0, 8))

        ctk.CTkButton(footer, text="⏻  Cerrar sesión", font=("Helvetica", 12, "bold"),
                      height=40, corner_radius=10,
                      fg_color="#EF4444", hover_color="#DC2626",
                      text_color="white", command=self.cerrar_sesion).pack(fill="x")

    def cargar_sidebar(self):
        for w in self.lista_frame.winfo_children():
            w.destroy()
        chats = cargar_chats_db(self._usuario_id)
        for chat in chats:
            titulo   = (chat.get("titulo") or "Chat sin título").strip()
            fecha    = chat.get("fecha", "")
            chat_id  = chat.get("chat_id", "")
            texto_btn = "💬  " + titulo[:26] + "\n" + fecha
            btn = ctk.CTkButton(
                self.lista_frame,
                text=texto_btn,
                font=("Helvetica", 12),
                height=56,
                corner_radius=8,
                fg_color=C["sidebar_btn"],
                hover_color=C["sidebar_hover"],
                text_color="white",
                anchor="w",
                command=lambda cid=chat_id: self.cargar_chat(cid)
            )
            btn.pack(fill="x", pady=3, padx=4)

    def cargar_chat(self, chat_id):
        try:
            chats = cargar_chats_db(self._usuario_id)
            chat  = next((c for c in chats if c["chat_id"] == chat_id), None)
            if not chat:
                return
            mensajes = json.loads(chat["mensajes"]) if isinstance(chat["mensajes"], str) else chat["mensajes"]
            self.historial = mensajes
            self._chat_archivo = chat_id
            self._chat_titulo  = chat.get("titulo", "")
            self._limpiar_chat()
            for m in self.historial:
                if m["role"] == "system":
                    continue
                contenido   = m["content"]
                imagen_ruta = m.get("imagen_ruta", None)
                if isinstance(contenido, list):
                    texto = next((p["text"] for p in contenido if p.get("type") == "text"), "📎 Imagen adjunta")
                    texto = (texto or "📎 Imagen adjunta") + "  🖼️"
                else:
                    texto = contenido
                BurbujaChat(self.chat_frame, texto, es_ia=(m["role"] == "assistant"),
                            avatar_ia=self.avatar_ia, imagen_ruta=imagen_ruta)
            self.chat_frame.update_idletasks()
            self.after(50, self._scroll_arriba)
        except Exception as e:
            print(f"[Chat] Error cargando: {e}")

    def _construir_main(self):
        self.main = ctk.CTkFrame(self, fg_color=C["bg_app"])
        self.main.grid(row=0, column=1, sticky="nsew")
        self.main.grid_rowconfigure(1, weight=1)
        self.main.grid_columnconfigure(0, weight=1)
        self._construir_header()
        self._construir_area_chat()
        self._construir_panel_video()
        self._construir_entrada()

    def _construir_header(self):
        self._header = ctk.CTkFrame(self.main, height=64, corner_radius=0,
                              fg_color=C["header_bg"], border_width=1, border_color=C["border"])
        header = self._header
        header.grid(row=0, column=0, columnspan=2, sticky="ew")
        header.grid_propagate(False)

        # Izquierda: solo título
        ctk.CTkLabel(header, text="Soporte de Incidencias IT", font=("Helvetica", 16, "bold"),
                     text_color=C["text_dark"]).pack(side="left", padx=16)

        # Derecha: info usuario + "En línea" + botones juntos
        nombre = self.usuario_data.get("nombre_completo", "")
        depto  = self.usuario_data.get("departamento", "")

        # Grupo de botones de icono pegados (☰ y ⚙ juntos)
        btn_group = ctk.CTkFrame(header, fg_color="transparent")
        btn_group.pack(side="right", padx=(0, 12))

        self.btn_prefs = ctk.CTkButton(btn_group, text="⚙", width=32, height=32,
            fg_color="transparent", hover_color=C["bg_app"], text_color=C["text_dark"],
            font=("Helvetica", 16), corner_radius=6, command=self._abrir_personalizacion)
        self.btn_prefs.pack(side="left", padx=2)

        self.btn_toggle = ctk.CTkButton(btn_group, text="☰", width=32, height=32,
            fg_color="transparent", hover_color=C["bg_app"], text_color=C["text_dark"],
            font=("Helvetica", 16), corner_radius=6, command=self.toggle_sidebar)
        self.btn_toggle.pack(side="left", padx=2)

        self._ventana_prefs = None

        # Info usuario
        ctk.CTkLabel(header, text=f"👤  {nombre}  |  {depto}",
                     font=("Helvetica", 11), text_color=C["text_muted"]).pack(side="right", padx=(0, 12))
        ctk.CTkLabel(header, text="● En línea", font=("Helvetica", 11),
                     text_color="#22C55E").pack(side="right", padx=(0, 16))

    def _construir_area_chat(self):
        self.chat_frame = ctk.CTkScrollableFrame(self.main, fg_color=C["bg_app"], scrollbar_button_color=C["border"])
        self.chat_frame.grid(row=1, column=0, sticky="nsew")
        self.welcome = ctk.CTkFrame(self.chat_frame, fg_color=C["card"], corner_radius=16,
                                    border_width=1, border_color=C["border"])
        self.welcome.pack(fill="x", padx=24, pady=32)
        ctk.CTkLabel(self.welcome, text="👋  Bienvenido a DigiHelp AI",
                     font=("Helvetica", 20, "bold"), text_color=C["text_dark"]).pack(pady=(24, 8))
        ctk.CTkLabel(self.welcome, text="Describe tu incidencia IT y te ayudaré a resolverla paso a paso.",
                     font=("Helvetica", 13), text_color=C["text_muted"]).pack(pady=(0, 8))
        sugerencias = ["Mi impresora no conecta a la red", "No puedo acceder a la VPN",
                       "Mi equipo está muy lento", "Olvidé mi contraseña de dominio"]
        grid = ctk.CTkFrame(self.welcome, fg_color="transparent")
        grid.pack(pady=(8, 24), padx=24)
        for i, s in enumerate(sugerencias):
            btn = ctk.CTkButton(grid, text=s, height=36, corner_radius=18,
                fg_color=C["bg_app"], hover_color=C["border"], text_color=C["accent"],
                border_width=1, border_color=C["accent"], font=("Helvetica", 12),
                command=lambda t=s: self._enviar_sugerencia(t))
            btn.grid(row=i // 2, column=i % 2, padx=6, pady=4, sticky="ew")
        self.avatar_ia = self._crear_avatar()

    def _construir_panel_video(self):
        # Panel lateral eliminado — ahora el video abre en ventana independiente
        self.video_label = None  # se crea dinámicamente en la ventana

    def _construir_entrada(self):
        self._entrada_bg = ctk.CTkFrame(self.main, corner_radius=0,
                                  fg_color=C["header_bg"], border_width=1, border_color=C["border"])
        self._entrada_bg.grid(row=2, column=0, columnspan=2, sticky="ew")

        # Preview de imagen adjunta (oculto por defecto)
        self._preview_frame = ctk.CTkFrame(self._entrada_bg, fg_color="#EFF6FF",
                                           corner_radius=8, border_width=1, border_color=C["accent"])
        self._lbl_preview = ctk.CTkLabel(self._preview_frame, text="",
                                         font=("Helvetica", 12), text_color=C["accent"])
        self._lbl_preview.pack(side="left", padx=10, pady=6)
        ctk.CTkButton(self._preview_frame, text="✕", width=24, height=24,
                      fg_color="transparent", hover_color="#DBEAFE",
                      text_color=C["accent"], font=("Helvetica", 12),
                      command=self._quitar_imagen).pack(side="right", padx=6)

        inner = ctk.CTkFrame(self._entrada_bg, fg_color="transparent", height=70)
        inner.pack(fill="x", expand=True, padx=20, pady=10)
        inner.pack_propagate(False)

        # Botón único de adjuntar
        self.btn_adjuntar = ctk.CTkButton(inner, text="📎", width=46, height=46,
            corner_radius=23, fg_color=C["bg_app"], hover_color=C["border"],
            text_color=C["text_dark"], font=("Helvetica", 18), border_width=1,
            border_color=C["border"], command=self._toggle_panel_adjuntar)
        self.btn_adjuntar.pack(side="left", padx=(0, 10))

        self._panel_adjuntar = None  # se crea como Toplevel al pulsar el botón

        self.entry = ctk.CTkEntry(inner, placeholder_text="Describe tu incidencia IT aquí...",
            height=46, corner_radius=23, fg_color=C["input_bg"], border_color=C["border"],
            border_width=1, font=("Helvetica", FUENTE), text_color=C["text_dark"])
        self.entry.pack(side="left", fill="x", expand=True, padx=(0, 12))
        self.entry.bind("<Return>", lambda e: self.lanzar_hilo())

        self.btn_enviar = ctk.CTkButton(inner, text="Enviar  ➤", width=110, height=46,
            corner_radius=23, fg_color=C["accent"], hover_color=C["accent_dark"],
            font=("Helvetica", 13, "bold"), text_color="white", command=self.lanzar_hilo)
        self.btn_enviar.pack(side="right")

    def _toggle_panel_adjuntar(self):
        if self._panel_adjuntar and self._panel_adjuntar.winfo_exists():
            self._cerrar_panel_adjuntar()
            return

        # Calcular posición encima del botón usando coordenadas absolutas
        self.update_idletasks()
        bx = self.btn_adjuntar.winfo_rootx()
        by = self.btn_adjuntar.winfo_rooty()

        # Crear ventana popup sin bordes y con esquinas redondeadas
        import tkinter as tk
        popup = tk.Toplevel(self)
        popup.overrideredirect(True)
        popup.attributes("-topmost", True)
        popup.configure(bg="#000001")  # color transparente
        try:
            popup.attributes("-transparentcolor", "#000001")  # Windows: esquinas redondeadas
        except Exception:
            popup.configure(bg=C["card"])
        popup.resizable(False, False)

        # Frame con esquinas redondeadas visibles
        outer = ctk.CTkFrame(popup, fg_color=C["card"], corner_radius=14, border_width=0)
        outer.pack(fill="both", expand=True, padx=6, pady=6)

        ctk.CTkLabel(outer, text="Adjuntar archivo",
            font=("Helvetica", 11, "bold"), text_color=C["text_muted"]).pack(
            pady=(12, 6), padx=16, anchor="w")

        ctk.CTkButton(outer, text="  🖼️   Imagen",
            height=40, corner_radius=8, anchor="w", width=200,
            fg_color="transparent", hover_color=C["bg_app"],
            text_color=C["text_dark"], font=("Helvetica", 13),
            command=lambda: [self._cerrar_panel_adjuntar(), self._adjuntar_imagen()]
        ).pack(fill="x", padx=6, pady=2)

        ctk.CTkButton(outer, text="  📄   Documento",
            height=40, corner_radius=8, anchor="w", width=200,
            fg_color="transparent", hover_color=C["bg_app"],
            text_color=C["text_dark"], font=("Helvetica", 13),
            command=lambda: [self._cerrar_panel_adjuntar(), self._adjuntar_documento()]
        ).pack(fill="x", padx=6, pady=(0, 10))

        # Posicionar encima del botón
        popup.update_idletasks()
        ph = popup.winfo_reqheight()
        pw = popup.winfo_reqwidth()
        popup.geometry(f"{pw}x{ph}+{bx}+{by - ph - 6}")

        self._panel_adjuntar = popup
        self.btn_adjuntar.configure(fg_color="#DBEAFE", border_color=C["accent"])

        # Cerrar al hacer click en cualquier sitio de la app principal
        self.bind("<Button-1>", self._click_fuera_panel, add="+")
        self.bind_all("<Escape>", lambda e: self._cerrar_panel_adjuntar())

    def _click_fuera_panel(self, event):
        self.after(50, self._cerrar_panel_adjuntar)

    def _cerrar_panel_adjuntar(self):
        if self._panel_adjuntar and self._panel_adjuntar.winfo_exists():
            self._panel_adjuntar.destroy()
        self._panel_adjuntar = None
        self.btn_adjuntar.configure(fg_color=C["bg_app"], border_color=C["border"])
        try:
            self.unbind("<Button-1>")
            self.unbind("<Escape>")
        except Exception:
            pass

    def _adjuntar_documento(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar documento",
            filetypes=[
                ("Documentos", "*.pdf *.docx *.xlsx *.txt"),
                ("PDF", "*.pdf"), ("Word", "*.docx"),
                ("Excel", "*.xlsx"), ("Texto", "*.txt"),
                ("Todos", "*.*")
            ]
        )
        if not ruta:
            return
        try:
            texto = self._extraer_texto_doc(ruta)
            if not texto.strip():
                raise ValueError("El documento está vacío o no se pudo leer.")
            # Truncar a 8000 caracteres para no saturar el contexto
            if len(texto) > 8000:
                texto = texto[:8000] + "\n\n[... documento truncado ...]"
            self._doc_adjunto = (ruta, texto)
            nombre = os.path.basename(ruta)
            self._lbl_preview.configure(text=f"📄  {nombre}")
            self._preview_frame.pack(fill="x", padx=20, pady=(8, 0),
                                     before=self._entrada_bg.winfo_children()[1])
            self.btn_adjuntar.configure(fg_color="#DBEAFE", border_color=C["accent"])
        except Exception as e:
            self._lbl_preview.configure(text=f"⚠️ Error: {e}")
            self._preview_frame.pack(fill="x", padx=20, pady=(8, 0),
                                     before=self._entrada_bg.winfo_children()[1])

    def _extraer_texto_doc(self, ruta):
        ext = ruta.lower().split(".")[-1]
        if ext == "pdf":
            doc = fitz.open(ruta)
            return "\n".join(page.get_text() for page in doc)
        elif ext == "docx":
            doc = DocxDocument(ruta)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        elif ext == "xlsx":
            wb = openpyxl.load_workbook(ruta, data_only=True)
            lineas = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lineas.append(f"--- Hoja: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    fila = [str(c) if c is not None else "" for c in row]
                    if any(fila):
                        lineas.append("\t".join(fila))
            return "\n".join(lineas)
        elif ext == "txt":
            with open(ruta, encoding="utf-8", errors="ignore") as f:
                return f.read()
        return ""

    def _quitar_documento(self):
        self._doc_adjunto = None
        self.btn_adjuntar.configure(fg_color=C["bg_app"], border_color=C["border"])

    def _mostrar_error_imagen(self, err):
        BurbujaChat(self.chat_frame, f"⚠️ No se pudo procesar la imagen: {err}",
                    es_ia=True, avatar_ia=self.avatar_ia)
        self._scroll_abajo()

    def _adjuntar_imagen(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar imagen",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.gif *.webp"), ("Todos", "*.*")]
        )
        if not ruta:
            return
        # Solo guardar la ruta — el base64 se genera en el hilo secundario al enviar
        ext = ruta.lower().split(".")[-1]
        media_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                      "png": "image/png", "gif": "image/gif",
                      "webp": "image/webp"}.get(ext, "image/png")
        self._imagen_adjunta = (ruta, None, media_type)  # base64 pendiente
        nombre = os.path.basename(ruta)
        self._lbl_preview.configure(text=f"📎  {nombre}")
        self._preview_frame.pack(fill="x", padx=20, pady=(8, 0), before=self._entrada_bg.winfo_children()[1])
        self.btn_adjuntar.configure(fg_color="#DBEAFE", border_color=C["accent"])

    def _quitar_imagen(self):
        self._imagen_adjunta = None
        self.btn_adjuntar.configure(fg_color=C["bg_app"], border_color=C["border"])
        if not self._doc_adjunto:
            self._preview_frame.pack_forget()

    def _enviar_sugerencia(self, texto):
        self.entry.delete(0, "end")
        self.entry.insert(0, texto)
        self.lanzar_hilo()

    def lanzar_hilo(self):
        msg = self.entry.get().strip()
        if not msg and not self._imagen_adjunta:
            return
        if self.welcome and self.welcome.winfo_exists():
            self.welcome.destroy()
            self.welcome = None
        self.entry.delete(0, "end")
        self.btn_enviar.configure(state="disabled", text="Enviando…")
        ts = datetime.now().strftime("%H:%M")

        # Si estamos esperando confirmación de resolución
        if self._esperando_confirmacion:
            msg_lower = msg.lower()
            # Detectar primero frases negativas (tienen prioridad)
            no_resuelto = any(p in msg_lower for p in [
                "no", "sigue", "todavía", "aún", "nada", "no funciona",
                "no va", "igual", "mismo", "sin funcionar", "tampoco",
                "sigue igual", "no me va", "no sirve", "no ha funcionado",
                "no funciona", "mal", "peor", "falla"
            ])
            # Solo marcar como resuelto si NO hay palabras negativas
            resuelto = not no_resuelto and any(p in msg_lower for p in [
                "sí", "si", "funciona", "resuelto", "gracias", "perfecto",
                "ya va", "solucionado", "ok", "vale", "genial", "listo",
                "ya funciona", "perfecto", "bien"
            ])

            if resuelto:
                # Problema resuelto — resetear todo
                self._intentos = 0
                self._esperando_confirmacion = False
                self._problema_inicial = None
                BurbujaChat(self.chat_frame, "¡Perfecto! Me alegra que se haya resuelto. Si tienes cualquier otra incidencia, aquí estaré. 😊",
                            es_ia=True, avatar_ia=self.avatar_ia, timestamp=ts)
                self._scroll_abajo()
                self.btn_enviar.configure(state="normal", text="Enviar  ➤")
                return

            elif no_resuelto:
                self._intentos += 1
                self._esperando_confirmacion = False

                if self._intentos >= 5:
                    # Máximo de intentos alcanzado — mostrar diálogo y NO continuar
                    BurbujaChat(self.chat_frame,
                                "He agotado mis intentos para resolver esta incidencia. Voy a crear un ticket para que el equipo de IT te ayude directamente.",
                                es_ia=True, avatar_ia=self.avatar_ia, timestamp=ts)
                    self._scroll_abajo()
                    self.after(800, self._mostrar_dialogo_incidencia)
                    self.btn_enviar.configure(state="normal", text="Enviar  ➤")
                    return
                # Si no llegó a 5, continuar con la IA normalmente

        imagen = self._imagen_adjunta
        doc    = self._doc_adjunto
        self._quitar_imagen()
        if doc:
            self._quitar_documento()
            self._preview_frame.pack_forget()

        # Mostrar burbuja del usuario
        if imagen and doc:
            texto_burbuja = (msg or "📎 Adjunto") + "  🖼️📄"
        elif imagen:
            texto_burbuja = (msg or "📎 Imagen adjunta") + "  🖼️"
        elif doc:
            nombre_doc = os.path.basename(doc[0])
            texto_burbuja = (msg or f"📄 {nombre_doc}") + f"  📄"
        else:
            texto_burbuja = msg
        ruta_img = imagen[0] if imagen else None
        BurbujaChat(self.chat_frame, texto_burbuja, es_ia=False, timestamp=ts, imagen_ruta=ruta_img, animar=True)
        self._scroll_abajo()

        es_primero = len([m for m in self.historial if m["role"] == "user"]) == 0
        # Pasar imagen como tupla (ruta, None, media_type) — base64 se genera en el hilo
        threading.Thread(
            target=self._responder_ia,
            args=(msg, es_primero, ts, imagen, doc),
            daemon=True
        ).start()

    def _fijar_titulo(self, texto_usuario):
        titulo = self._generar_titulo(texto_usuario)
        self._chat_titulo  = titulo
        if not self._chat_archivo:
            self._chat_archivo = self.chat_id

    def _responder_ia(self, texto_usuario, es_primero, ts, imagen=None, doc=None):
        # Fijar archivo desde el principio para todos los mensajes del chat
        if not self._chat_archivo:
            self._chat_archivo = self.chat_id

        # Codificar imagen a base64 aquí (hilo secundario, no bloquea la UI)
        usa_vision = False
        if imagen:
            try:
                ruta, _, media_type = imagen
                with open(ruta, "rb") as f:
                    b64 = base64.standard_b64encode(f.read()).decode("utf-8")
                contenido_user = [
                    {"type": "text", "text": texto_usuario or "Analiza esta imagen y ayúdame a resolver el problema de IT que ves."},
                    {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{b64}"}}
                ]
                usa_vision = True
            except Exception as e:
                contenido_user = texto_usuario or ""
                self.after(0, lambda: self._mostrar_error_imagen(str(e)))
        else:
            contenido_user = texto_usuario or ""

        # Si hay documento adjunto, añadir su contenido al mensaje
        if doc:
            _, texto_doc = doc
            nombre_doc = os.path.basename(doc[0])
            if isinstance(contenido_user, list):
                contenido_user[0]["text"] += f"\n\n---\nDocumento adjunto: {nombre_doc}\n{texto_doc}"
            else:
                contenido_user = f"{contenido_user}\n\n---\nDocumento adjunto: {nombre_doc}\n{texto_doc}".strip()

        # AÑADIR MENSAJE DEL USUARIO AL HISTORIAL
        entrada = {"role": "user", "content": contenido_user}
        if imagen:
            entrada["imagen_ruta"] = imagen[0]
        if doc:
            entrada["doc_nombre"] = os.path.basename(doc[0])
        self.historial.append(entrada)

        # Comprobar si hay video con mensaje predefinido
        ruta_v, msg_v = self._buscar_video(texto_usuario)
        if ruta_v and msg_v:
            self.historial.append({"role": "assistant", "content": msg_v})
            self.after(0, lambda m=msg_v: BurbujaChat(
                self.chat_frame, m, es_ia=True,
                avatar_ia=self.avatar_ia, timestamp=ts))
            self.after(0, self._scroll_abajo)
            self.after(300, lambda r=ruta_v: self.reproducir_video(r))
            if es_primero:
                self._fijar_titulo(texto_usuario)
            self._guardar_chat()
            self.after(0, self.cargar_sidebar)
            self.after(0, lambda: self.btn_enviar.configure(state="normal", text="Enviar  ➤"))
            return

        # Respuesta normal de la IA
        burbuja_ref = [None]
        creada = threading.Event()

        def crear_burbuja():
            burbuja_ref[0] = BurbujaChat(self.chat_frame, "⏳ Escribiendo",
                                         es_ia=True, avatar_ia=self.avatar_ia, timestamp=ts, animar=True)
            burbuja_ref[0].iniciar_puntos()
            self._scroll_abajo()
            creada.set()

        self.after(0, crear_burbuja)
        creada.wait(timeout=3.0)

        respuesta = ""
        try:
            # Usar visión si el mensaje actual o alguno anterior tiene imagen
            hay_imagen_en_historial = any(
                isinstance(m["content"], list) for m in self.historial
            )
            modelo_actual = MODELO_VISION if (usa_vision or hay_imagen_en_historial) else MODELO_TEXTO

            # Limpiar historial: quitar imagen_ruta pero mantener imágenes en formato API
            historial_limpio = []
            for m in self.historial:
                historial_limpio.append({"role": m["role"], "content": m["content"]})
            stream = client.chat.completions.create(
                model=modelo_actual, messages=historial_limpio,
                max_tokens=1024, stream=True, timeout=30)
            for chunk in stream:
                delta = chunk.choices[0].delta.content
                if delta:
                    respuesta += delta
                    snap = respuesta
                    if burbuja_ref[0]:
                        if hasattr(burbuja_ref[0], '_puntos_activos') and burbuja_ref[0]._puntos_activos:
                            burbuja_ref[0].detener_puntos()
                        self.after(0, lambda t=snap: burbuja_ref[0].actualizar_texto(t))
                    self.after(0, self._scroll_abajo)

            self.historial.append({"role": "assistant", "content": respuesta})

            if es_primero:
                self._fijar_titulo(texto_usuario)
                self._problema_inicial = texto_usuario
            self._guardar_chat()
            self.after(0, self.cargar_sidebar)

            if ruta_v:
                self.after(200, lambda r=ruta_v: self.reproducir_video(r))

            # Activar espera de confirmación tras cada respuesta IT
            if respuesta.strip() and not (ruta_v and msg_v):
                self._esperando_confirmacion = True

        except Exception as e:
            err = str(e)
            if burbuja_ref[0]:
                self.after(0, lambda: burbuja_ref[0].actualizar_texto(f"⚠️ Error: {err}"))
        finally:
            self.after(0, lambda: self.btn_enviar.configure(state="normal", text="Enviar  ➤"))

    def _mostrar_dialogo_incidencia(self):
        """Crea el ticket automáticamente con los datos del usuario logueado."""
        nombre    = self.usuario_data.get("nombre_completo", "Usuario desconocido")
        depto     = self.usuario_data.get("departamento", "")
        email     = self.usuario_data.get("email", "")
        problema  = self._problema_inicial or "Sin descripción"
        urgencia  = detectar_urgencia(problema)

        exito = crear_incidencia_db(nombre, depto, email, problema, urgencia)
        ts    = datetime.now().strftime("%H:%M")
        sep   = "\n"

        if exito:
            urgencia_emoji = {"alta": "🔴", "media": "🟡", "baja": "🟢"}.get(urgencia, "⚪")
            msg_ticket = (
                f"✅ He creado un ticket automáticamente con tus datos:{sep}"
                f"👤 Usuario: {nombre}  |  🏢 Departamento: {depto}{sep}"
                f"📧 Email: {email}{sep}"
                f"🔧 Problema: {problema}{sep}"
                f"{urgencia_emoji} Urgencia detectada: {urgencia.upper()}{sep}"
                f"El equipo de IT revisará tu incidencia lo antes posible."
            )
        else:
            msg_ticket = "⚠️ No se pudo crear el ticket. Por favor contacta directamente con el equipo de IT."

        # Añadir al historial y mostrar en el chat
        self.historial.append({"role": "assistant", "content": msg_ticket})
        self.after(0, lambda m=msg_ticket: BurbujaChat(
            self.chat_frame, m, es_ia=True, avatar_ia=self.avatar_ia, timestamp=ts))
        self.after(0, self._scroll_abajo)
        self._guardar_chat()
        self.after(0, self.cargar_sidebar)
        self._intentos = 0
        self._problema_inicial = None
        self._esperando_confirmacion = False

    def _guardar_chat(self):
        if not self._chat_archivo:
            self._chat_archivo = self.chat_id
        if not self._chat_titulo:
            self._chat_titulo = self._chat_archivo

        # Limpiar mensajes con imagen antes de guardar
        mensajes_limpios = []
        for m in self.historial:
            if isinstance(m["content"], list):
                texto = next((p["text"] for p in m["content"] if p.get("type") == "text"), "")
                ruta_img = m.get("imagen_ruta", "")
                mensajes_limpios.append({
                    "role": m["role"],
                    "content": (texto or "📎 Imagen adjunta") + "  🖼️",
                    "imagen_ruta": ruta_img
                })
            else:
                mensajes_limpios.append(m)

        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        mensajes_json = json.dumps(mensajes_limpios, ensure_ascii=False)
        guardar_chat_db(self._usuario_id, self._chat_archivo, self._chat_titulo, fecha, mensajes_json)

    def _generar_titulo(self, msg):
        try:
            r = client.chat.completions.create(model=MODELO_TEXTO, max_tokens=20,
                messages=[{"role": "user", "content": f"Resume en 4 palabras máximo (solo el título, sin comillas ni puntos): {msg}"}])
            titulo = r.choices[0].message.content.strip().strip('"').strip("'").strip(".")
            return titulo if titulo else "Incidencia IT"
        except Exception:
            return "Incidencia IT"

    def _buscar_video(self, texto):
        """
        Busca un video por palabra clave.
        El campo ruta_video puede ser:
          - Un ID de Google Drive (solo el ID, ej: 1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs)
          - Una URL de Drive (https://drive.google.com/file/d/ID/view)
          - Una ruta local (C:\\videos\\video.mp4 o relativa)
        """
        if not texto:
            return None, None
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT palabra_clave, ruta_video, mensaje FROM videos")
            rows = cursor.fetchall()
            cursor.close()
            conn.close()
            texto_lower = texto.lower()
            for row in rows:
                if row["palabra_clave"].lower() in texto_lower:
                    ruta = row.get("ruta_video", "") or ""
                    drive_id = _extraer_drive_id(ruta)
                    if drive_id:
                        # ── Es un video de Google Drive: descargar a temp ──
                        burbuja_descarga = [None]
                        ts = datetime.now().strftime("%H:%M")
                        self.after(0, lambda: self._mostrar_descargando(burbuja_descarga, ts))
                        ruta_temp = descargar_video_drive(drive_id)
                        self.after(0, lambda b=burbuja_descarga: self._ocultar_descargando(b))
                        if ruta_temp:
                            self._video_temp = ruta_temp
                            return ruta_temp, row.get("mensaje")
                        else:
                            return None, None
                    else:
                        # ── Ruta local (comportamiento original) ──
                        return ruta, row.get("mensaje")
        except Exception as e:
            print(f"[DB] Error: {e}")
        return None, None

    def _mostrar_descargando(self, ref, ts):
        ref[0] = BurbujaChat(self.chat_frame,
                             "⏳ Descargando tutorial desde Drive…",
                             es_ia=True, avatar_ia=self.avatar_ia, timestamp=ts)
        self._scroll_abajo()

    def _ocultar_descargando(self, ref):
        try:
            if ref[0]:
                ref[0].destroy()
        except Exception:
            pass

    def reproducir_video(self, ruta_video):
        # Si es una ruta temporal (descargada de Drive) ya es absoluta y existe
        if not os.path.isabs(ruta_video):
            ruta_video = recurso_path(ruta_video)
        if not os.path.exists(ruta_video):
            print(f"[Video] No encontrado: {ruta_video}")
            return
        self.cerrar_video(borrar_temp=False)  # no borramos el temp que acabamos de descargar

        # ── Instancia VLC ─────────────────────────────
        self._vlc_instance = vlc.Instance("--no-xlib")
        self._vlc_player   = self._vlc_instance.media_player_new()
        media = self._vlc_instance.media_new(ruta_video)
        self._vlc_player.set_media(media)
        self.video_activo = True

        # ── Ventana ───────────────────────────────────
        self._ventana_video = ctk.CTkToplevel(self)
        self._ventana_video.title("DigiHelp — Tutorial")
        self._ventana_video.geometry("760x520")
        self._ventana_video.configure(fg_color=C["video_bg"])
        self._ventana_video.attributes("-topmost", True)
        self._ventana_video.protocol("WM_DELETE_WINDOW", self.cerrar_video)
        self._ventana_video.resizable(False, False)

        # ── Header ────────────────────────────────────
        header_v = ctk.CTkFrame(self._ventana_video, fg_color="#1E293B", height=48)
        header_v.pack(fill="x")
        header_v.pack_propagate(False)
        ctk.CTkLabel(header_v, text="▶  Tutorial DigiHelp",
                     font=("Helvetica", 14, "bold"), text_color="white").pack(side="left", padx=16)
        ctk.CTkButton(header_v, text="✕ Cerrar", width=80, height=30,
                      fg_color="#EF4444", hover_color="#DC2626", text_color="white",
                      font=("Helvetica", 12), command=self.cerrar_video).pack(side="right", padx=12, pady=8)

        # ── Canvas donde VLC renderiza el vídeo ───────
        import tkinter as tk
        self._video_canvas = tk.Frame(self._ventana_video, bg="black")
        self._video_canvas.pack(expand=True, fill="both")

        # ── Controles ─────────────────────────────────
        controles = ctk.CTkFrame(self._ventana_video, fg_color="#1E293B", height=100)
        controles.pack(fill="x")
        controles.pack_propagate(False)

        # Slider de progreso
        self._progress_var = ctk.DoubleVar(value=0)
        self._slider = ctk.CTkSlider(
            controles, from_=0, to=1000,
            variable=self._progress_var,
            button_color=C["accent"], button_hover_color=C["accent_dark"],
            progress_color=C["accent"], fg_color="#374151", height=14,
            command=self._vlc_saltar
        )
        self._slider.pack(fill="x", padx=16, pady=(10, 2))

        # Tiempo
        self._lbl_tiempo = ctk.CTkLabel(controles, text="0:00 / 0:00",
                                        font=("Helvetica", 11), text_color=C["text_muted"])
        self._lbl_tiempo.pack()

        # Botones
        btns = ctk.CTkFrame(controles, fg_color="transparent")
        btns.pack(pady=(2, 6))

        estilo_btn = dict(width=52, height=36, corner_radius=10,
                          fg_color="#374151", hover_color="#4B5563",
                          text_color="white", font=("Helvetica", 16))

        ctk.CTkButton(btns, text="⏮", **estilo_btn,
                      command=lambda: self._vlc_saltar_segundos(-30)).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="⏪", **estilo_btn,
                      command=lambda: self._vlc_saltar_segundos(-10)).pack(side="left", padx=4)

        self._btn_play = ctk.CTkButton(btns, text="⏸", width=60, height=36,
                      corner_radius=10, fg_color=C["accent"], hover_color=C["accent_dark"],
                      text_color="white", font=("Helvetica", 18),
                      command=self._vlc_toggle_pausa)
        self._btn_play.pack(side="left", padx=8)

        ctk.CTkButton(btns, text="⏩", **estilo_btn,
                      command=lambda: self._vlc_saltar_segundos(10)).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="⏭", **estilo_btn,
                      command=lambda: self._vlc_saltar_segundos(30)).pack(side="left", padx=4)

        # Volumen
        vol_frame = ctk.CTkFrame(controles, fg_color="transparent")
        vol_frame.place(relx=1.0, rely=0.1, anchor="ne", x=-16)
        ctk.CTkLabel(vol_frame, text="🔊", font=("Helvetica", 13),
                     text_color=C["text_muted"]).pack(side="left", padx=(0, 4))
        self._vol_slider = ctk.CTkSlider(vol_frame, from_=0, to=100, width=90, height=14,
                      button_color=C["accent"], button_hover_color=C["accent_dark"],
                      progress_color=C["accent"], fg_color="#374151",
                      command=lambda v: self._vlc_player.audio_set_volume(int(v)))
        self._vol_slider.set(80)
        self._vol_slider.pack(side="left")

        # ── Asignar canvas a VLC tras renderizar ──────
        def _embed_vlc():
            self._ventana_video.update_idletasks()
            hwnd = self._video_canvas.winfo_id()
            import sys
            if sys.platform == "win32":
                self._vlc_player.set_hwnd(hwnd)
            elif sys.platform == "darwin":
                self._vlc_player.set_nsobject(hwnd)
            else:
                self._vlc_player.set_xwindow(hwnd)
            self._vlc_player.audio_set_volume(80)
            self._vlc_player.play()
            self._vlc_actualizar_ui()

        self._ventana_video.after(200, _embed_vlc)

    def _vlc_toggle_pausa(self):
        if self._vlc_player:
            self._vlc_player.pause()
            pausado = self._vlc_player.get_state() == vlc.State.Paused
            self._btn_play.configure(text="▶" if pausado else "⏸")

    def _vlc_saltar_segundos(self, segundos):
        if self._vlc_player:
            t = self._vlc_player.get_time() + segundos * 1000
            t = max(0, t)
            self._vlc_player.set_time(int(t))

    def _vlc_saltar(self, valor):
        if self._vlc_player:
            dur = self._vlc_player.get_length()
            if dur > 0:
                self._vlc_player.set_time(int(float(valor) / 1000 * dur))

    def _vlc_actualizar_ui(self):
        if not self.video_activo or not self._vlc_player:
            return
        try:
            if not self._ventana_video.winfo_exists():
                return
            t   = self._vlc_player.get_time()
            dur = self._vlc_player.get_length()
            if dur > 0:
                self._progress_var.set(t / dur * 1000)
                sa = t // 1000;  sd = dur // 1000
                self._lbl_tiempo.configure(
                    text=f"{sa // 60}:{sa % 60:02d} / {sd // 60}:{sd % 60:02d}")
            # Detectar fin
            if self._vlc_player.get_state() == vlc.State.Ended:
                self._vlc_player.stop()
                self._vlc_player.set_time(0)
                self._btn_play.configure(text="▶")
                return
            self._ventana_video.after(500, self._vlc_actualizar_ui)
        except Exception:
            pass

    def cerrar_video(self, borrar_temp=True):
        self.video_activo = False
        if self._vlc_player:
            try:
                self._vlc_player.stop()
                self._vlc_player.release()
            except Exception:
                pass
            self._vlc_player = None
        if self._vlc_instance:
            try:
                self._vlc_instance.release()
            except Exception:
                pass
            self._vlc_instance = None
        if hasattr(self, "_ventana_video") and self._ventana_video.winfo_exists():
            self._ventana_video.destroy()
        # Borrar archivo temporal descargado de Drive
        if borrar_temp and hasattr(self, "_video_temp") and self._video_temp:
            try:
                os.unlink(self._video_temp)
                print(f"[Drive] Temp eliminado: {self._video_temp}")
            except Exception:
                pass
            self._video_temp = None
        # Compatibilidad con cerrar_sesion
        self.video_cap = None

    def _crear_avatar(self):
        try:
            img = Image.open(recurso_path("avatar.png")).convert("RGBA")
            s = min(img.size)
            img = img.crop(((img.width-s)//2, (img.height-s)//2, (img.width+s)//2, (img.height+s)//2))
            img = img.resize((36, 36), Image.LANCZOS)
            mask = Image.new("L", (36, 36), 0)
            ImageDraw.Draw(mask).ellipse((0, 0, 36, 36), fill=255)
            out = Image.new("RGBA", (36, 36), (0, 0, 0, 0))
            out.paste(img, (0, 0), mask)
            return ctk.CTkImage(light_image=out, dark_image=out, size=(36, 36))
        except Exception:
            return None

    def _limpiar_chat(self):
        for w in self.chat_frame.winfo_children():
            w.destroy()
        self.welcome = None

    def _scroll_arriba(self):
        self.chat_frame._parent_canvas.yview_moveto(0.0)

    def _scroll_abajo(self):
        self.chat_frame._parent_canvas.yview_moveto(1.0)

    def nuevo_chat(self):
        self.chat_id = f"chat_{int(time.time())}"
        self._chat_titulo = None
        self._chat_archivo = None
        self._imagen_adjunta = None
        self._doc_adjunto = None
        self._intentos = 0            # Contador de intentos de resolución
        self._problema_inicial = None # Primer mensaje del usuario
        self._esperando_confirmacion = False  # Esperando si se resolvió
        self.historial = [{"role": "system", "content": SYSTEM_PROMPT}]
        self._limpiar_chat()
        self.cerrar_video()
        self.welcome = ctk.CTkFrame(self.chat_frame, fg_color=C["card"], corner_radius=16,
                                    border_width=1, border_color=C["border"])
        self.welcome.pack(fill="x", padx=24, pady=32)
        ctk.CTkLabel(self.welcome, text="👋  Bienvenido a DigiHelp AI",
                     font=("Helvetica", 20, "bold"), text_color=C["text_dark"]).pack(pady=(24, 8))
        ctk.CTkLabel(self.welcome, text="Describe tu incidencia IT y te ayudaré a resolverla paso a paso.",
                     font=("Helvetica", 13), text_color=C["text_muted"]).pack(pady=(0, 24))

    def _abrir_personalizacion(self):
        if self._ventana_prefs and self._ventana_prefs.winfo_exists():
            self._ventana_prefs.focus()
            return

        prefs = cargar_prefs()
        win = ctk.CTkToplevel(self)
        self._ventana_prefs = win
        win.title("Personalización")
        win.geometry("400x460")
        win.resizable(False, False)
        win.configure(fg_color=C["bg_app"])
        win.attributes("-topmost", True)
        win.grab_set()

        # Centrar respecto a la app principal
        self.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width()  - 400) // 2
        y = self.winfo_rooty() + (self.winfo_height() - 460) // 2
        win.geometry(f"400x460+{x}+{y}")

        # ── Header ventana ──
        header = ctk.CTkFrame(win, fg_color=C["sidebar_bg"], corner_radius=0, height=56)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="⚙️  Personalización", font=("Helvetica", 15, "bold"),
                     text_color="white").pack(side="left", padx=20, pady=16)

        body = ctk.CTkFrame(win, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=24, pady=16)

        # ── Tema ──
        ctk.CTkLabel(body, text="Tema", font=("Helvetica", 13, "bold"),
                     text_color=C["text_dark"], anchor="w").pack(fill="x", pady=(0, 6))
        tema_var = ctk.StringVar(value=prefs.get("tema", "claro"))
        temas_frame = ctk.CTkFrame(body, fg_color="transparent")
        temas_frame.pack(fill="x", pady=(0, 16))
        for valor, etiqueta in [("claro", "☀️  Claro"), ("oscuro", "🌙  Oscuro")]:
            ctk.CTkRadioButton(temas_frame, text=etiqueta, variable=tema_var, value=valor,
                font=("Helvetica", 13), text_color=C["text_dark"],
                fg_color=C["accent"], hover_color=C["accent_dark"]).pack(side="left", padx=(0, 20))

        # Separador
        ctk.CTkFrame(body, fg_color=C["border"], height=1).pack(fill="x", pady=(0, 16))

        # ── Color de acento ──
        ctk.CTkLabel(body, text="Color de acento", font=("Helvetica", 13, "bold"),
                     text_color=C["text_dark"], anchor="w").pack(fill="x", pady=(0, 10))
        acento_var = ctk.StringVar(value=prefs.get("acento", "Azul"))
        acentos_frame = ctk.CTkFrame(body, fg_color="transparent")
        acentos_frame.pack(fill="x", pady=(0, 16))
        for c in range(3):
            acentos_frame.grid_columnconfigure(c, weight=1)

        btns_acento = {}
        def _sel_acento(nombre):
            acento_var.set(nombre)
            for n, b in btns_acento.items():
                color, _ = ACENTOS[n]
                b.configure(border_width=3 if n == nombre else 0,
                            border_color="white" if n == nombre else color)

        for i, (nombre, (color, dark)) in enumerate(ACENTOS.items()):
            fila, col = divmod(i, 3)
            btn = ctk.CTkButton(acentos_frame, text=nombre, height=34,
                corner_radius=8, fg_color=color, hover_color=dark,
                text_color="white", font=("Helvetica", 12, "bold"),
                command=lambda n=nombre: _sel_acento(n))
            btn.grid(row=fila, column=col, padx=4, pady=3, sticky="ew")
            btns_acento[nombre] = btn
        _sel_acento(acento_var.get())

        # Separador
        ctk.CTkFrame(body, fg_color=C["border"], height=1).pack(fill="x", pady=(0, 16))

        # ── Tamaño de fuente ──
        ctk.CTkLabel(body, text="Tamaño de fuente", font=("Helvetica", 13, "bold"),
                     text_color=C["text_dark"], anchor="w").pack(fill="x", pady=(0, 6))
        fuente_var = ctk.IntVar(value=prefs.get("fuente", 13))
        fuente_row = ctk.CTkFrame(body, fg_color="transparent")
        fuente_row.pack(fill="x", pady=(0, 20))
        lbl_fuente = ctk.CTkLabel(fuente_row, text=f"{fuente_var.get()} px",
                                   font=("Helvetica", 12), text_color=C["text_muted"], width=50)
        lbl_fuente.pack(side="right")
        def _on_fuente(val):
            lbl_fuente.configure(text=f"{int(float(val))} px")
        ctk.CTkSlider(fuente_row, from_=11, to=17, number_of_steps=6,
                      variable=fuente_var, fg_color=C["border"],
                      button_color=C["accent"], button_hover_color=C["accent_dark"],
                      progress_color=C["accent"], command=_on_fuente).pack(side="left", fill="x", expand=True, padx=(0, 10))

        # ── Botones ──
        btns_row = ctk.CTkFrame(body, fg_color="transparent")
        btns_row.pack(fill="x", pady=(0, 4))

        def _aplicar():
            nuevas = {"tema": tema_var.get(), "acento": acento_var.get(), "fuente": int(fuente_var.get())}
            guardar_prefs(nuevas)                          # fallback local
            guardar_prefs_db(self._usuario_id, nuevas)    # DB por usuario
            aplicar_tema(nuevas)
            win.destroy()
            self._ventana_prefs = None
            # Aplicar colores a widgets existentes sin destruir el chat
            self.configure(fg_color=C["bg_app"])
            self.sidebar.configure(fg_color=C["sidebar_bg"])
            self._icono_frame.configure(fg_color=C["accent"])
            self.btn_nuevo.configure(fg_color=C["accent"], hover_color=C["accent_dark"])
            self.lista_frame.configure(scrollbar_button_color=C["sidebar_hover"])
            self._header.configure(fg_color=C["header_bg"], border_color=C["border"])
            self.btn_toggle.configure(hover_color=C["bg_app"], text_color=C["text_dark"])
            self.btn_prefs.configure(hover_color=C["bg_app"], text_color=C["text_dark"])
            self.chat_frame.configure(fg_color=C["bg_app"])
            self.main.configure(fg_color=C["bg_app"])
            self._entrada_bg.configure(fg_color=C["header_bg"], border_color=C["border"])
            self.btn_enviar.configure(fg_color=C["accent"], hover_color=C["accent_dark"])
            self.btn_adjuntar.configure(hover_color=C["border"])
            self.entry.configure(border_color=C["border"], font=("Helvetica", FUENTE))
            # Recolorear burbujas existentes en el chat actual
            for widget in self.chat_frame.winfo_children():
                if isinstance(widget, BurbujaChat):
                    widget.recolorear()
            # Recargar sidebar para que los botones de chats cojan el nuevo color
            self.cargar_sidebar()

        ctk.CTkButton(btns_row, text="Aplicar", height=40, corner_radius=10,
                      fg_color=C["accent"], hover_color=C["accent_dark"],
                      font=("Helvetica", 13, "bold"), text_color="white",
                      command=_aplicar).pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(btns_row, text="Cancelar", height=40, corner_radius=10,
                      fg_color=C["border"], hover_color=C["text_muted"],
                      font=("Helvetica", 13), text_color=C["text_dark"],
                      command=win.destroy).pack(side="left", fill="x", expand=True)

    def cerrar_sesion(self):
        self.cerrar_video()
        self.destroy()
        login = LoginApp()
        login.mainloop()

    def toggle_sidebar(self):
        if self.sidebar_visible:
            self.sidebar.grid_forget()
            self.sidebar_visible = False
        else:
            self.sidebar.grid(row=0, column=0, sticky="nsew")
            self.sidebar_visible = True

if __name__ == "__main__":
    login = LoginApp()
    login.mainloop()